"""Estilos corporativos compartidos para exportaciones Excel (openpyxl)."""
from time_utils import ahora_mexico

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

CORPORATE = {
    'navy': '0F2847',
    'navy_mid': '1E3A5F',
    'gold': 'B49A5A',
    'header_text': 'FFFFFF',
    'alt_row': 'F1F5F9',
    'border': 'CBD5E1',
    'text': '1E293B',
    'muted': '64748B',
    'brand': 'Internet SP',
    'dept': 'Departamento Tecnico',
}


def _thin_border():
    side = Side(style='thin', color=CORPORATE['border'])
    return Border(left=side, right=side, top=side, bottom=side)


def _ajustar_anchos_columnas(ws, fila_inicio_datos, fila_encabezado):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 12
        for row_idx in range(fila_encabezado, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, min(len(str(val)), 48))
        ws.column_dimensions[letter].width = max_len + 2
    ws.freeze_panes = ws.cell(row=fila_inicio_datos + 1, column=1).coordinate


def exportar_dataframe_corporativo(writer, df, sheet_name, titulo_reporte, subtitulo=None):
    """Escribe un DataFrame con cabecera corporativa en la hoja indicada."""
    fila_encabezado = 6
    fila_inicio_datos = fila_encabezado + 1
    subtitulo = subtitulo or f"Generado el {ahora_mexico().strftime('%d/%m/%Y %H:%M')}"

    df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=fila_encabezado - 1)

    ws = writer.sheets[sheet_name]
    max_col = max(len(df.columns), 1)
    last_letter = get_column_letter(max_col)

    fill_navy = PatternFill('solid', fgColor=CORPORATE['navy'])
    fill_gold = PatternFill('solid', fgColor=CORPORATE['gold'])
    fill_header = PatternFill('solid', fgColor=CORPORATE['navy'])
    fill_alt = PatternFill('solid', fgColor=CORPORATE['alt_row'])
    font_title = Font(bold=True, size=14, color=CORPORATE['header_text'])
    font_sub = Font(size=10, color=CORPORATE['muted'])
    font_brand = Font(bold=True, size=11, color=CORPORATE['navy'])
    font_th = Font(bold=True, size=9, color=CORPORATE['header_text'])
    font_td = Font(size=9, color=CORPORATE['text'])
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(vertical='center', wrap_text=True)

    ws.merge_cells(f'A1:{last_letter}1')
    c1 = ws['A1']
    c1.value = CORPORATE['brand']
    c1.font = font_title
    c1.fill = fill_navy
    c1.alignment = align_center

    ws.merge_cells(f'A2:{last_letter}2')
    c2 = ws['A2']
    c2.value = titulo_reporte
    c2.font = Font(bold=True, size=12, color=CORPORATE['navy'])
    c2.alignment = align_center

    ws.merge_cells(f'A3:{last_letter}3')
    c3 = ws['A3']
    c3.value = subtitulo
    c3.font = font_sub
    c3.alignment = align_center

    ws.merge_cells(f'A4:{last_letter}4')
    ws['A4'].fill = fill_gold
    ws.row_dimensions[4].height = 4

    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=fila_encabezado, column=col_idx)
        cell.font = font_th
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = _thin_border()

    for row_idx in range(fila_inicio_datos, ws.max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = font_td
            cell.alignment = align_left
            cell.border = _thin_border()
            if (row_idx - fila_inicio_datos) % 2 == 1:
                cell.fill = fill_alt

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 16
    ws.row_dimensions[fila_encabezado].height = 20

    _ajustar_anchos_columnas(ws, fila_inicio_datos, fila_encabezado)


def crear_workbook_corporativo_multihoja(dataframes_por_hoja, titulo_documento):
    """
    dataframes_por_hoja: lista de dicts {sheet_name, df, titulo}
    Retorna BytesIO listo para send_file.
    """
    from io import BytesIO
    import pandas as pd

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for item in dataframes_por_hoja:
            exportar_dataframe_corporativo(
                writer,
                item['df'],
                item['sheet_name'],
                item.get('titulo', item['sheet_name']),
                item.get('subtitulo', titulo_documento),
            )
    output.seek(0)
    return output
